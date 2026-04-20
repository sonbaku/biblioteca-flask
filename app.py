from flask import Flask, render_template, request, redirect, Response
import psycopg2
import os
from pymongo import MongoClient
from bson import ObjectId
import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl

app = Flask(__name__)

# POSTGRESQL
conexion = psycopg2.connect(
    host="dpg-d7ipq54vikkc73enaad0-a.oregon-postgres.render.com",
    database="biblioteca_db2",
    user="biblioteca_db2_user",
    password="ZL1gCzsnurlVvdROL4ZcdLkOoMtIjzLS",
    port="5432",
    sslmode="require"
)
cursor = conexion.cursor()

# MONGODB
MONGO_URI = os.environ.get("MONGO_URI", "mongodb://localhost:27017/")
MONGO_DB  = os.environ.get("MONGO_DB", "biblioteca")
mongo_client  = MongoClient(MONGO_URI)
mongo_db      = mongo_client[MONGO_DB]
col_productos = mongo_db["productos"]


@app.route('/')
def login():
    return render_template("login.html")


@app.route('/login', methods=['POST'])
def validar_login():
    usuario  = request.form['usuario']
    password = request.form['password']
    cursor.execute(
        "SELECT * FROM usuarios WHERE usuario=%s AND password=%s",
        (usuario, password)
    )
    user = cursor.fetchone()
    if user:
        return redirect('/menu')
    else:
        return render_template("login.html", error="Usuario o contraseña incorrectos")


@app.route('/menu')
def menu():
    return render_template("menu.html")


@app.route('/libros')
def libros():
    cursor.execute("SELECT * FROM libros")
    libros = cursor.fetchall()
    return render_template("libros.html", libros=libros)


@app.route('/agregar')
def agregar():
    return render_template("agregar.html")


@app.route('/guardar', methods=['POST'])
def guardar():
    titulo = request.form['titulo']
    autor  = request.form['autor']
    anio   = int(request.form['anio'])
    precio = float(request.form['precio'])
    if anio < 1000 or anio > 2100:
        return "Error: Año inválido"
    if precio < 0 or precio > 1000000:
        return "Error: Precio inválido"
    try:
        cursor.execute(
            "INSERT INTO libros (titulo, autor, anio, precio) VALUES (%s,%s,%s,%s)",
            (titulo, autor, anio, precio)
        )
        conexion.commit()
    except:
        conexion.rollback()
        return "Error al guardar el libro"
    return redirect('/libros')


@app.route('/eliminar/<int:id>')
def eliminar(id):
    try:
        cursor.execute("DELETE FROM libros WHERE id = %s", (id,))
        conexion.commit()
    except:
        conexion.rollback()
        return "Error al eliminar"
    return redirect('/libros')


@app.route('/editar/<int:id>')
def editar(id):
    cursor.execute("SELECT * FROM libros WHERE id = %s", (id,))
    libro = cursor.fetchone()
    return render_template("editar.html", libro=libro)


@app.route('/actualizar/<int:id>', methods=['POST'])
def actualizar(id):
    titulo = request.form['titulo']
    autor  = request.form['autor']
    anio   = int(request.form['anio'])
    precio = float(request.form['precio'])
    if anio < 1000 or anio > 2100:
        return "Error: Año inválido"
    if precio < 0 or precio > 1000000:
        return "Error: Precio inválido"
    try:
        cursor.execute(
            "UPDATE libros SET titulo=%s, autor=%s, anio=%s, precio=%s WHERE id=%s",
            (titulo, autor, anio, precio, id)
        )
        conexion.commit()
    except:
        conexion.rollback()
        return "Error al actualizar el libro"
    return redirect('/libros')


# PRODUCTOS MONGODB
@app.route('/productos')
def productos():
    lista = list(col_productos.find())
    return render_template("productos.html", productos=lista)


@app.route('/producto/nuevo')
def producto_nuevo():
    return render_template("producto_agregar.html")


@app.route('/producto/guardar', methods=['POST'])
def producto_guardar():
    nombre    = request.form['nombre']
    categoria = request.form['categoria']
    precio    = float(request.form['precio'])
    stock     = int(request.form['stock'])
    col_productos.insert_one({
        "nombre": nombre,
        "categoria": categoria,
        "precio": precio,
        "stock": stock
    })
    return redirect('/productos')


@app.route('/producto/editar/<id>')
def producto_editar(id):
    producto = col_productos.find_one({"_id": ObjectId(id)})
    return render_template("producto_editar.html", producto=producto)


@app.route('/producto/actualizar/<id>', methods=['POST'])
def producto_actualizar(id):
    nombre    = request.form['nombre']
    categoria = request.form['categoria']
    precio    = float(request.form['precio'])
    stock     = int(request.form['stock'])
    col_productos.update_one(
        {"_id": ObjectId(id)},
        {"$set": {"nombre": nombre, "categoria": categoria,
                  "precio": precio, "stock": stock}}
    )
    return redirect('/productos')


@app.route('/producto/eliminar/<id>')
def producto_eliminar(id):
    col_productos.delete_one({"_id": ObjectId(id)})
    return redirect('/productos')


# EXPORTAR PDF
@app.route('/productos/exportar/pdf')
def exportar_pdf():
    buffer  = io.BytesIO()
    doc     = SimpleDocTemplate(buffer, pagesize=letter)
    estilos = getSampleStyleSheet()
    elementos = []
    elementos.append(Paragraph("Stock de Productos", estilos['Title']))
    elementos.append(Paragraph(" ", estilos['Normal']))
    datos = [["Nombre", "Categoría", "Precio", "Stock"]]
    for p in col_productos.find():
        datos.append([
            str(p.get("nombre", "")),
            str(p.get("categoria", "")),
            f"${p.get('precio', 0):.2f}",
            str(p.get("stock", 0))
        ])
    tabla = Table(datos, colWidths=[160, 140, 100, 80])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 12),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F8')]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN',      (2, 0), (-1, -1), 'CENTER'),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={"Content-Disposition": "attachment; filename=productos.pdf"}
    )


# EXPORTAR XLSX
@app.route('/productos/exportar/xlsx')
def exportar_xlsx():
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(["Nombre", "Categoría", "Precio", "Stock"])
    azul = PatternFill("solid", fgColor="2E86AB")
    for celda in ws[1]:
        celda.font      = Font(color="FFFFFF", bold=True)
        celda.fill      = azul
        celda.alignment = Alignment(horizontal="center")
    for p in col_productos.find():
        ws.append([
            p.get("nombre", ""),
            p.get("categoria", ""),
            p.get("precio", 0),
            p.get("stock", 0)
        ])
    for col in ws.columns:
        max_len = max(len(str(celda.value or "")) for celda in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=productos.xlsx"}
    )


@app.route('/logout')
def logout():
    return redirect('/')


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)